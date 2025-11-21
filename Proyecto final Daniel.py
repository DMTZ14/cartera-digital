import openpyxl
import time
from time import asctime,localtime
wb = openpyxl.load_workbook("CarteraDigital.xlsx")
wb.active = 0
Hoja = wb.active
def FechaHoy():
    texto = asctime(localtime())
    tiempo = texto.split(" ")
    meses = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
    }
    dia = tiempo[2]
    mes = meses[tiempo[1]]
    anio = tiempo[4]
    fecha = f"{dia}/{mes}/{anio}"
    return fecha


def mostrar_datos():
    wb.active = 0
    Hoja = wb.active
    for r in range(1, Hoja.max_row + 1):
        for c in range(1, Hoja.max_column + 1):
            print(format(Hoja.cell(row=r, column=c).value, "13"), end='\t\t')
        print()
    print()
if True:
    gasto_1=""
    ingreso_1=""
    cash=0
    necesidad=0
    ocio=0

    ahorro=0
    otros=0
    saldo=0
    presupuesto_semanal=0
    gasto_mensual=0
    presupuesto_mensual=0
    gasto_semanal=0
while True:
    print("Menú Cartera Digital")
    print("1. Consultar movimientos.")
    print("2. Registrar movimientos.")
    print("3. Presupuestos")
    print("4. Salir")
    menu=int(input("Escoge una opción: "))

    while menu==1:
        print("")
        print("---Consultar Finanzas---")
        print("1. Consultar Saldo.")
        print("2. Consultar Movimientos.")
        print("3. Consultar gastos por categorías.")
        print("4. Consultar mes actual")
        print("5. Consultar semana actual")
        print("6. Consultar gráficas de gastos.")
        print("7. Consultar gráficas de ingresos.")
        print("8. Salir")
        print("")
        menu_int= int(input("Escoge una opción: "))
        if menu_int==1:
            print()
            print("---Consultar Saldo---")
            saldo=0
            try:
                for i in range(2, Hoja.max_row + 1):
                    saldo+=Hoja.cell(row=i, column=4).value
                print(f"Saldo total (Ingresos-Gastos): {saldo}")
            except:
                print(f"Saldo total: $ 0.00")
            print()
        elif menu_int==2:
            print()
            print("---Consultar Movimientos---")
            print()
            mostrar_datos()
            print()
        elif menu_int==3:
            print()
            print("---Consultar gastos por categoría---")
            necesidad=0
            ocio=0
            ahorro=0
            otros=0
            gastos=0
            for i in range(2, Hoja.max_row + 1):
                if Hoja.cell(i,3).value=="Gasto":
                    gastos -= Hoja.cell(row=i, column=4).value


            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value=="Necesidad":
                    necesidad+=Hoja.cell(row=i, column=4).value
            print(f"Necesidad: {necesidad}, ({-(necesidad/gastos):.1%})")

            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value=="Ocio":
                    ocio+=Hoja.cell(row=i, column=4).value
            print(f"Ocio: {ocio}, ({-(ocio/gastos):.1%})")

            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value=="Ahorro":
                    ahorro+=Hoja.cell(row=i, column=4).value
            print(f"Ahorro: {ahorro}, ({-(ahorro/gastos):.1%})")

            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value=="Otros":
                    otros+=Hoja.cell(row=i, column=4).value
            print(f"Otros: {otros}, ({-(otros/gastos):.1%})")
            print()

        elif menu_int==4:
            print()
            print("---Consultar mes actual---")
            mes_actual = int(time.strftime("%m", time.localtime()))
            anio_actual = int(time.strftime("%Y", time.localtime()))

            print(f"Movimientos del mes {mes_actual}/{anio_actual}:")
            for c in range(1, Hoja.max_column + 1):
                print(format(Hoja.cell(row=1, column=c).value, "10"), end='\t\t')
            print()


            for r in range(1, Hoja.max_row + 1):
                fecha_celda = Hoja.cell(row=r, column=2).value
                try:
                    dia, mes, anio = fecha_celda.split("/")
                    if int(mes) == mes_actual and int(anio) == anio_actual:
                        for c in range(1, Hoja.max_column + 1):
                            print(format(Hoja.cell(r, column=c).value,"8"), end='\t\t')
                        print()
                except:

                    continue
        elif menu_int==5:
            print()
            print("---Consultar semana actual---")
            semana_actual = int(time.strftime("%W", time.localtime()))
            anio_actual = time.localtime().tm_year

            print(f"Movimientos de la semana {semana_actual} del año {anio_actual}:")
            for c in range(1, Hoja.max_column + 1):
                print(format(Hoja.cell(row=1, column=c).value, "10"), end='\t\t')
            print()
            for i in range(1, Hoja.max_row + 1):
                try:
                    semana_celda = Hoja.cell(row=i, column=7).value
                    if semana_celda == semana_actual:
                        for c in range(1, Hoja.max_column + 1):
                            print(format(Hoja.cell(row=i, column=c).value,"10"), end='\t\t')
                        print()
                except:
                    continue

        elif menu_int==6:
            print()
            print("---Consultar gráficas de gasto---")
            gastos=0
            necesidad=0
            ocio=0
            ahorro=0
            otros=0

            for i in range(2, Hoja.max_row + 1):
                if Hoja.cell(i,3).value=="Gasto":
                    gastos += Hoja.cell(row=i, column=4).value
            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value == "Necesidad":
                    necesidad += Hoja.cell(row=i, column=4).value
            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value == "Ocio":
                    ocio += Hoja.cell(row=i, column=4).value
            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value == "Ahorro":
                    ahorro += Hoja.cell(row=i, column=4).value
            for i in range(1, Hoja.max_row + 1):
                if Hoja.cell(row=i, column=5).value == "Otros":
                    otros += Hoja.cell(row=i, column=4).value

            categorias={"Necesidad":-necesidad,"Ocio":-ocio,"Ahorro":-ahorro,"Otros":-otros}
            for i in categorias:
                print(format(i,"10")+"|",end="")
                print(format(categorias[i],"7")+"|", end="")
                print(round((categorias[i]/-gastos)*10)*"||||||||||")
            print()

        elif menu_int==7:
            print()
            print("---Consultar gráficas de ingreso---")
            ingresos = 0
            sueldo = 0
            venta = 0
            regalo = 0
            ahorro_retirado = 0
            otro = 0
            for i in range(2, Hoja.max_row + 1):
                if Hoja.cell(i, 3).value == "Ingreso":
                    ingresos += Hoja.cell(i, 4).value
            for i in range(2, Hoja.max_row + 1):
                if Hoja.cell(i, 3).value == "Ingreso":
                    categoria = Hoja.cell(i, 5).value
                    monto = Hoja.cell(i, 4).value
                    if categoria == "Sueldo":
                        sueldo += monto
                    elif categoria == "Venta":
                        venta += monto
                    elif categoria == "Regalo":
                        regalo += monto
                    elif categoria == "Ahorro retirado":
                        ahorro_retirado += monto
                    elif categoria == "Otro":
                        otro += monto
            categorias = {"Sueldo": sueldo,"Venta": venta,"Regalo": regalo,"Ahorro retirado": ahorro_retirado,"Otro": otro}

            for i in categorias:
                print(format(i, "15") + "|", end="")
                print(format(categorias[i], "7") + "|", end="")
                print(round((categorias[i] / ingresos) * 10) * "||||||||||")

            print()

        elif menu_int==8:
            print()
            print("Salir al menú principal.")
            print()
            break
        continue

    while menu==2:
        print("")
        print("---Registrar movimientos---")
        print("1. Registrar Gasto.")
        print("2. Registrar Ingreso.")
        print("3. Editar movimiento.")
        print("4. Eliminar movimiento.")
        print("5. Salir al menú principal")

        menu_int = int(input("Escoge una opción: "))
        if menu_int==1:
            otro_gasto=1
            while otro_gasto==1:
                print("")
                print("---Registrar Gasto---")
                wb.active = 0
                Hoja = wb.active
                Fecha = FechaHoy()
                Tipo = "Gasto"
                Monto = -float(input("Monto: "))
                print("¿De qué categoría fue tu gasto?")
                print("1. Necesidad")
                print("2. Ocio")
                print("3. Ahorro")
                print("4. Otros")
                categoria_gasto = int(input("Escoge una opción:"))
                while categoria_gasto > 4 or categoria_gasto < 1:
                    print("Opción no válida.")
                    categoria_gasto = int(input("Escoge una opción:"))
                if categoria_gasto == 1:
                    Categoria = "Necesidad"
                elif categoria_gasto == 2:
                    Categoria = "Ocio"
                elif categoria_gasto == 3:
                    Categoria = "Ahorro"
                elif categoria_gasto == 4:
                    Categoria = "Otros"
                Concepto = input("Concepto: ")
                semana_actual = int(time.strftime("%W", time.localtime()))
                SiguienteInsert = Hoja.max_row
                try:
                    Hoja.cell(row=(Hoja.max_row + 1), column=1).value = SiguienteInsert
                    Hoja.cell(row=(Hoja.max_row), column=2).value = Fecha
                    Hoja.cell(row=(Hoja.max_row), column=3).value = Tipo
                    Hoja.cell(row=(Hoja.max_row), column=4).value = Monto
                    Hoja.cell(row=(Hoja.max_row), column=5).value = Categoria
                    Hoja.cell(row=(Hoja.max_row), column=6).value = Concepto
                    Hoja.cell(row=(Hoja.max_row), column=7).value = semana_actual

                    wb.save("CarteraDigital.xlsx")
                    print("Insertado Exitosamente")
                except:
                    print("Ocurrió un error al escribir")
                print()

                print("Quisieras registras otro gasto? \n1. Sí\n2. No")
                otro_gasto=int(input("Escoge una opción: "))


        elif menu_int==2:
            print("")
            otro_ingreso = 1
            while otro_ingreso == 1:
                print("")
                print("---Registrar Ingreso---")
                wb.active = 0
                Hoja = wb.active
                Fecha = FechaHoy()
                Tipo = "Ingreso"
                Monto = float(input("Monto: "))
                print("¿De qué categoría es tu ingreso?")
                print("1. Sueldo")
                print("2. Venta")
                print("3. Regalo")
                print("4. Ahorro retirado")
                print("5. Otro")
                categoria_ingreso = int(input("Escoge una opción:"))
                if categoria_ingreso == 1:
                    Categoria = "Sueldo"
                elif categoria_ingreso == 2:
                    Categoria = "Venta"
                elif categoria_ingreso == 3:
                    Categoria = "Regalo"
                elif categoria_ingreso == 4:
                    Categoria = "Ahorro retirado"
                elif categoria_ingreso == 5:
                    Categoria = "Otro"
                Concepto = input("Concepto: ")
                semana_actual = int(time.strftime("%W", time.localtime()))
                SiguienteInsert = Hoja.max_row
                try:
                    Hoja.cell(row=(Hoja.max_row + 1), column=1).value = SiguienteInsert
                    Hoja.cell(row=(Hoja.max_row), column=2).value = Fecha
                    Hoja.cell(row=(Hoja.max_row), column=3).value = Tipo
                    Hoja.cell(row=(Hoja.max_row), column=4).value = Monto
                    Hoja.cell(row=(Hoja.max_row), column=5).value = Categoria
                    Hoja.cell(row=(Hoja.max_row), column=6).value = Concepto
                    Hoja.cell(row=(Hoja.max_row), column=7).value = semana_actual
                    wb.save("CarteraDigital.xlsx")
                    print("Insertado Exitosamente")
                except:
                    print("Ocurrió un error al escribir")

                print("Quisieras registras otro ingreso? \n1. Sí\n2. No")
                otro_ingreso = int(input("Escoge una opción: "))

        elif menu_int==3:
            print()
            print("---Editar Movimiento---")

            mostrar_datos()
            movimiento_editar= int(input("¿Que movimiento deseas Editar?: "))+1

            #Fecha = ""
            #Tipo = ""
            Monto = float(input("Monto: "))
            if Hoja.cell(row=(movimiento_editar), column=3).value=="Gasto":
                Monto=-Monto
                print("¿De qué categoría fue tu gasto?")
                print("1. Necesidad")
                print("2. Ocio")
                print("3. Ahorro")
                print("4. Otros")
                categoria_gasto = int(input("Escoge una opción:"))
                while categoria_gasto > 4 or categoria_gasto < 1:
                    print("Opción no válida.")
                    categoria_gasto = int(input("Escoge una opción:"))
                if categoria_gasto == 1:
                    necesidad += Monto
                    Categoria = "Necesidad"
                elif categoria_gasto == 2:
                    ocio += Monto
                    Categoria = "Ocio"
                elif categoria_gasto == 3:
                    ahorro += Monto
                    Categoria = "Ahorro"
                elif categoria_gasto == 4:
                    otros += Monto
                    Categoria = "Otros"
                Hoja.cell(row=(movimiento_editar), column=5).value = Categoria
            Concepto = input("Concepto: ")
            try:
                Hoja.cell(row=(movimiento_editar), column=1).value = movimiento_editar-1
                #Hoja.cell(row=(movimiento_editar), column=2).value = Fecha
                #Hoja.cell(row=(movimiento_editar), column=3).value = Tipo
                Hoja.cell(row=(movimiento_editar), column=4).value = Monto
                #Hoja.cell(row=(movimiento_editar), column=5).value = Categoria
                Hoja.cell(row=(movimiento_editar), column=6).value = Concepto
                wb.save("CarteraDigital.xlsx")
                print("Editado Exitosamente")
                print()
            except:
                print("Ocurrió un error al escribir")



        elif menu_int==4:
            print()
            print("---Eliminar Movimiento---")
            mostrar_datos()

            try:
                movimiento_eliminar= int(input("¿Que movimiento deseas eliminar?: "))+1
                Hoja.delete_rows(movimiento_eliminar)
                for i in range(2,Hoja.max_row+1):
                    Hoja.cell(i,1).value=i-1
                print("¡Movimiento Eliminado!")
            except:
                print("Ocurió un error.")
            wb.save("CarteraDigital.xlsx")

        elif menu_int==5:
            print()
            print("---Salir al menú principal----")
            break

    while menu == 3:
        print()
        print("---Presupuestos---")
        print("1. Establecer presupuesto semanal")
        print("2. Establecer presupuesto mensual")
        print("3. Ver cuánto queda del presupuesto")
        print("4. Modificar presupuesto")
        print("5. Volver al menú principal")
        wb.active = 1
        Hoja = wb.active

        menu_int = int(input("Escoge una opción: "))

        if menu_int == 1:
            print()
            print("---Establecer presupuesto semanal---")
            presupuesto_semanal=int(input("¿Cuanto presupuesto tienes para gastar por semana?:"))
            Hoja.cell(1,2).value=presupuesto_semanal
            wb.save("CarteraDigital.xlsx")

        elif menu_int == 2:
            print()
            print("---Establecer presupuesto mensual---")
            presupuesto_mensual=int(input("¿Cuanto presupuesto tienes para gastar por mes?:"))
            Hoja.cell(2, 2).value = presupuesto_mensual
            wb.save("CarteraDigital.xlsx")


        elif menu_int == 3:
            print()
            print("---Ver cuánto queda del presupuesto---")

            if Hoja.cell(2, 2).value== None or Hoja.cell(1,2).value==None:
                print("Presupuestos no definidos.")
            else:
                presupuesto_semanal = Hoja.cell(1, 2).value
                presupuesto_mensual = Hoja.cell(2, 2).value
                gasto_semanal = 0
                gasto_mensual = 0

                wb.active = 0
                Hoja = wb.active
                semana_actual = int(time.strftime("%W", time.localtime()))
                for i in range(2, Hoja.max_row + 1):
                    if Hoja.cell(row=i, column=7).value==semana_actual and Hoja.cell(i,3).value=="Gasto":
                        gasto_semanal += Hoja.cell(row=i, column=4).value
                print(f"Presupuesto semanal: {presupuesto_semanal+gasto_semanal}")


                mes_actual = int(time.strftime("%m", time.localtime()))
                for i in range(2, Hoja.max_row + 1):
                    fecha_celda = Hoja.cell(row=i, column=2).value
                    dia, mes, anio = fecha_celda.split("/")
                    if int(mes) == mes_actual and Hoja.cell(i,3).value=="Gasto":
                        gasto_mensual+=Hoja.cell(row=i, column=4).value
                print(f"Presupuesto mensual: {presupuesto_mensual+gasto_mensual}")

        elif menu_int == 4:
            print()
            print("---Modificar presupuesto----")
            modificar_presupuesto=int(input("¿Deseas modificar algún presupuesto?:\n1. Presupuesto Semanal\n2. Presupuesto Mensual\n"))
            if modificar_presupuesto==1:
                presupuesto_semanal = int(input("¿Cuanto presupuesto tienes para gastar por semana?:"))
                Hoja.cell(1, 2).value = presupuesto_semanal
            elif modificar_presupuesto==2:
                presupuesto_mensual = int(input("¿Cuanto presupuesto tienes para gastar por mes?:"))
                Hoja.cell(2, 2).value = presupuesto_mensual

        elif menu_int==5:
            print()
            print("Salir a menú principal.")
            print()
            break

    if menu==4:
        wb.close()
        break

